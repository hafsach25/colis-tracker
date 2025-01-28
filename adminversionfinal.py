import tkinter as tk
from tkinter import messagebox, ttk
import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
from tkinter import filedialog
import openpyxl
import numpy as np


from datetime import datetime


from tkcalendar import DateEntry
style = ttk.Style()
style.configure(
    "TCombobox",
    background="#f0f0f0",
    fieldbackground="white",
    foreground="#333",
    arrowcolor="#333"
)

plt.style.use('ggplot')  # ou 'fivethirtyeight', 'bmh', 'classic'
plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Arial'],
    'font.size': 10,
    'axes.titlesize': 12,
    'axes.labelsize': 11,
    'figure.facecolor': 'white',
    'axes.facecolor': 'white',
    'axes.grid': True,
    'grid.alpha': 0.3
})

# Style pour ttk
style = ttk.Style()
style.configure(
    "TCombobox",
    background="#f0f0f0",
    fieldbackground="white",
    foreground="#333",
    arrowcolor="#333"
)

# Style pour Treeview
style.configure(
    "Treeview",
    background="white",
    fieldbackground="white",
    foreground="black",
    rowheight=30
)
style.configure(
    "Treeview.Heading",
    background="#f0f0f0",
    foreground="#333",
    font=('Arial', 10, 'bold')
)
style.map(
    "Treeview",
    background=[('selected', '#2196F3')],
    foreground=[('selected', 'white')]
)
class AdminPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)
        self.after_ids = []  # Ajoutez cette ligne
        self.configure(fg_color="white")
        

        # Couleurs
        self.orange = "#ff8c00"
        self.blue_ciel = "#87cefa"
        self.bg_color = "#f2f2f2"
        self.blue_ciel = "#2196F3"
        self.orange = "#FF9800"
        self.green = "#4CAF50"
        self.red = "#F44336"
        

        # Barre supérieure
        self.framesup = ctk.CTkFrame(self, height=70, corner_radius=0, fg_color=self.blue_ciel)
        self.framesup.pack(fill="x", side="top")
        self.speedy_label = ctk.CTkLabel(self.framesup, text="Administration", font=("Helvetica", 50, "bold"), text_color="#ff8c00")
        self.speedy_label.pack(side="left",padx=5,pady=(10, 30))

        # Agencement principal
        main_frame = ctk.CTkFrame(self, fg_color="#f2f2f2", bg_color="white")
        main_frame.pack(fill="both", expand=True)

        # Barre latérale
        sidebar_frame = ctk.CTkScrollableFrame(main_frame, fg_color=self.blue_ciel,width=300)
        sidebar_frame.pack(side="left", fill="y", padx=0, pady=10)

        # Contenu principal
        self.content_frame = ctk.CTkScrollableFrame(main_frame,fg_color="white")
        self.content_frame.pack(side="right", fill="both", expand=True, padx=10)

        # Boutons dans la barre latérale
        self.create_sidebar_button(sidebar_frame, "👥 Gestion des clients", self.manage_clients)
        self.create_sidebar_button(sidebar_frame, "📦 Suivi des colis", self.tracking_packages)
        self.create_sidebar_button(sidebar_frame, "📊 Statistiques", self.view_statistics)
        self.create_sidebar_button(sidebar_frame, "🔔 Envoi des notifications", self.send_notifications)
        self.create_sidebar_button(sidebar_frame, "👨‍💼 Gestion des employés", self.employee)
        self.create_sidebar_button(sidebar_frame, "🚚  Gestion des véhicules", self.manage_vehicles)
        self.create_sidebar_button(sidebar_frame, "🏠 Demandes Déménagement", self.manage_moving_requests)
        self.create_sidebar_button(sidebar_frame, "🚪 Déconnexion", self.logout)
        self.show_welcome()
    def show_welcome(self):
        # Nettoyer le contenu actuel
        self.clear_content()
        
        # Frame principal
        welcome_frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        welcome_frame.pack(fill="both", expand=True, padx=40, pady=40)
        
        # Message de bienvenue
        welcome_label = ctk.CTkLabel(
            welcome_frame,
            text="👋 Bienvenue dans votre Espace Administration",
            font=("Arial", 32, "bold"),
            text_color=self.orange
        )
        welcome_label.pack(pady=(0, 30))

        # Sous-titre
        subtitle = ctk.CTkLabel(
            welcome_frame,
            text="Guide d'utilisation des fonctionnalités",
            font=("Arial", 24, "bold"),
            text_color=self.blue_ciel
        )
        subtitle.pack(pady=(0, 20))

        # Guide des fonctionnalités
        features = [
            ("👥 Gestion des clients", "Gérez votre base de clients, ajoutez, modifiez ou supprimez des informations client"),
            ("📦 Suivi des colis", "Suivez en temps réel l'état et la position des colis en cours de livraison"),
            ("📊 Statistiques", "Visualisez les statistiques et analyses de performance"),
            ("🔔 Envoi des notifications", "Envoyez des notifications aux clients et aux livreurs"),
            ("👨‍💼 Gestion des employés", "Gérez les informations et les accès des employés"),
            ("🚚 Gestion des véhicules", "Suivez et gérez votre flotte de véhicules"),
            ("🏠 Demandes Déménagement", "Gérez les demandes de déménagement des clients"),
            ("🚪 Déconnexion", "Déconnectez-vous de votre session")
        ]

        # Créer un frame scrollable pour les fonctionnalités
        features_frame = ctk.CTkScrollableFrame(
            welcome_frame,
            fg_color="transparent",
            height=400
        )
        features_frame.pack(fill="x", pady=20)

        # Ajouter chaque fonctionnalité avec son explication
        for icon_title, description in features:
            feature_frame = ctk.CTkFrame(features_frame, fg_color=self.bg_color)
            feature_frame.pack(fill="x", pady=5, padx=10)
            
            title = ctk.CTkLabel(
                feature_frame,
                text=icon_title,
                font=("Arial", 16, "bold"),
                text_color=self.orange
            )
            title.pack(anchor="w", padx=10, pady=(5, 0))
            
            desc = ctk.CTkLabel(
                feature_frame,
                text=description,
                font=("Arial", 12),
                text_color="gray",
                wraplength=800
            )
            desc.pack(anchor="w", padx=10, pady=(0, 5))

        # Note de bas de page
        footer_note = ctk.CTkLabel(
            welcome_frame,
            text="Pour commencer, cliquez sur l'une des options du menu à gauche",
            font=("Arial", 14, "italic"),
            text_color="gray"
        )
        footer_note.pack(pady=20)
    def manage_moving_requests(self):
        self.clear_content()
        
        # Frame principal avec titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        
        # Titre
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Déménagements",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame pour la barre de recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche par ID
        search_label = ctk.CTkLabel(search_frame, text="Rechercher par ID:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        
        search_entry = ctk.CTkEntry(search_frame, width=100)
        search_entry.pack(side="left", padx=5)
        def search_by_id():
          """Rechercher une demande par ID"""
          search_id=search_entry.get()
          
          if not search_id:
            messagebox.showwarning("Attention", "Veuillez entrer un ID")
            return
            
          found = False
          for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            if str(values[0]) == search_id:
                # Sélectionner et mettre en évidence l'élément trouvé
                self.tree.selection_set(item)
                self.tree.see(item)
                found = True
                break
        
          if not found:
            messagebox.showwarning("Recherche", "Aucune demande trouvée avec cet ID")
        
        search_button = ctk.CTkButton(
            search_frame,
            text="Rechercher",
            command= search_by_id,
            width=100
        )
        search_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des déménagements
        columns = ("ID", "Client", "Adresse départ", "Adresse arrivée", "Date", "Volume", "Prix", "Statut")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.tree.heading(col, text=col)
            width = 150 if col in ["Client", "Adresse départ", "Adresse arrivée"] else 100
            self.tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement du tableau et scrollbar
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame séparé pour les boutons APRÈS le tableau
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons
        accept_button = ctk.CTkButton(
            button_frame,
            text="Accepter",
            command=self.change_status_to_accepted,
            fg_color="green",
            width=120,
            height=32
        )
        accept_button.pack(side="left", padx=5)

        reject_button = ctk.CTkButton(
            button_frame,
            text="Refuser",
            command=self.change_status_to_rejected,
            fg_color="red",
            width=120,
            height=32
        )
        reject_button.pack(side="left", padx=5)

        # Charger les données
        self.load_sample_datadem()
    
    def change_status_to_accepted(self):
        """Changer le statut à 'Accepté'"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Attention", "Veuillez sélectionner une demande")
            return
            
        for item in selected_items:
            values = list(self.tree.item(item)['values'])
            if values[7] != "Accepté":
                values[7] = "Accepté"
                self.tree.item(item, values=values)
                messagebox.showinfo("Succès", f"La demande de {values[1]} a été acceptée")

    def change_status_to_rejected(self):
        """Changer le statut à 'Refusé'"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Attention", "Veuillez sélectionner une demande")
            return
            
        for item in selected_items:
            values = list(self.tree.item(item)['values'])
            if values[7] != "Refusé":
                values[7] = "Refusé"
                self.tree.item(item, values=values)
                messagebox.showinfo("Succès", f"La demande de {values[1]} a été refusée")
    def load_sample_datadem(self):
        """Charger des données d'exemple"""
        sample_data = [
            (1, "Jean Dupont", "123 rue Paris", "456 rue Lyon", "2024-03-20", "30m³", "1500€", "En attente"),
            (2, "Marie Martin", "789 rue Marseille", "321 rue Lille", "2024-03-22", "25m³", "1200€", "Accepté"),
            (3, "Pierre Durant", "147 rue Bordeaux", "258 rue Toulouse", "2024-03-25", "40m³", "2000€", "Refusé"),
            (4, "Sophie Lefebvre", "369 rue Nice", "741 rue Nantes", "2024-03-28", "35m³", "1800€", "En attente")
        ]
        
        # Nettoyer le tableau avant d'ajouter les données
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Ajouter les nouvelles données
        for item in sample_data:
            self.tree.insert('', 'end', values=item)

    def filter_requests(self):
        """Filtrer les demandes selon la recherche et le statut"""
        search_term = self.search_var.get().lower()
        status = self.status_var.get()
        
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            show = True
            
            if search_term:
                show = any(search_term in str(value).lower() for value in values)
            
            if status != "Tous":
                show = show and values[7] == status
            
            if show:
                self.tree.reattach(item, '', 'end')
            else:
                self.tree.detach(item)

    def reset_filters(self):
        """Réinitialiser les filtres"""
        self.search_var.set("")
        self.status_var.set("Tous")
        self.load_sample_data()  # Recharger les données

    def sort_column(self, col):
        """Trier le tableau par colonne"""
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        l.sort(reverse=getattr(self, '_sort_reverse', False))
        self._sort_reverse = not getattr(self, '_sort_reverse', False)
        
        for index, (_, k) in enumerate(l):
            self.tree.move(k, '', index)

    def clear_content(self):
        # Annuler les "after" en cours
        for after_id in self.after_ids:
            self.after_cancel(after_id)
        self.after_ids.clear()
        
        # Supprimer les widgets
        for widget in self.content_frame.winfo_children():
            widget.destroy()


    def create_sidebar_button(self, frame, text, command):
        button = ctk.CTkButton(frame, text=text, font=("arial", 20, "bold"), command=command, width=55, height=60, corner_radius=8, fg_color=self.orange)
        button.pack(fill="x", pady=25)

    
    def manage_vehicles(self):
        self.clear_content()
    
    # Initialiser les variables avec self comme master
        self.vehicle_search_var = tk.StringVar(master=self)
        self.vehicle_state_var = tk.StringVar(master=self, value="Tous")
        
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Véhicules",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche
        search_label = ctk.CTkLabel(search_frame, text="Rechercher:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        self.vehicle_search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.vehicle_search_var,
            width=200,
            placeholder_text="Matricule..."
        )
        search_entry.pack(side="left", padx=5)

        # Filtre par état
        state_label = ctk.CTkLabel(search_frame, text="État:", font=("Arial", 12))
        state_label.pack(side="left", padx=5)
        
        self.vehicle_state_var = tk.StringVar(value="Tous")
        state_combo = ttk.Combobox(
            search_frame,
            textvariable=self.vehicle_state_var,
            values=["Tous", "Disponible", "En mission", "En maintenance"],
            width=15,
            state="readonly"
        )
        state_combo.pack(side="left", padx=5)
        def search_vehicle():
          search_term = search_entry.get().lower()
          state_filter = state_combo.get()

          for item in self.vehicle_tree.get_children():
            values = self.vehicle_tree.item(item)['values']
            show = True

            if search_term and not str(values[0]).lower().startswith(search_term):
                show = False
            
            if state_filter != "Tous" and values[4] != state_filter:
                show = False

            if show:
                self.vehicle_tree.reattach(item, '', 'end')
            else:
                self.vehicle_tree.detach(item)

        def reset_vehicle_search():
          search_entry.insert(0,"")
          state_combo.set("Tous")
          self.load_vehicle_data()

        # Boutons de recherche et réinitialisation
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_vehicle,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)

        reset_button = ctk.CTkButton(
            search_frame,
            text="↺ Réinitialiser",
            command=reset_vehicle_search,
            width=100,
            fg_color=self.orange
        )
        reset_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des véhicules
        columns = ("Matricule", "Marque", "Modèle", "Année", "État", "Kilométrage", "Dernière maintenance")
        self.vehicle_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.vehicle_tree.heading(col, text=col)
            width = 150 if col in ["Marque", "Modèle"] else 100
            self.vehicle_tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.vehicle_tree.yview)
        self.vehicle_tree.configure(yscrollcommand=scrollbar.set)
        
        # Placement du tableau et scrollbar
        self.vehicle_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame séparé pour les boutons APRÈS le tableau
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons d'action
        add_button = ctk.CTkButton(
            button_frame,
            text="➕ Ajouter",
            command=self.show_add_vehicle_form,
            fg_color="green",
            width=120,
            height=32
        )
        add_button.pack(side="left", padx=5)

        modify_button = ctk.CTkButton(
            button_frame,
            text="✏️ Modifier",
            command=self.show_modify_vehicle_form,
            fg_color="orange",
            width=120,
            height=32
        )
        modify_button.pack(side="left", padx=5)

        delete_button = ctk.CTkButton(
            button_frame,
            text="🗑️ Supprimer",
            command=self.delete_vehicle,
            fg_color="red",
            width=120,
            height=32
        )
        delete_button.pack(side="left", padx=5)
        export_button = ctk.CTkButton(
        button_frame,
        text="📊 Exporter Excel",
        command=self.export_to_excelveh,
        fg_color="#2E7D32",  # Vert foncé
        width=120,
        height=32
    )
        export_button.pack(side="left", padx=5)


    # Charger les données
        self.load_vehicle_data()


    def export_to_excelveh(self):
        try:
            # Récupérer toutes les données du tableau
            data = []
            columns = ["Matricule", "Marque", "Modèle", "Année", "État", "Kilométrage", "Dernière maintenance"]
            
            for item in self.vehicle_tree.get_children():
                values = self.vehicle_tree.item(item)['values']
                data.append(values)

            # Créer un DataFrame pandas
            df = pd.DataFrame(data, columns=columns)

            # Demander à l'utilisateur où sauvegarder le fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Enregistrer le fichier Excel"
            )

            if file_path:
                # Créer un writer Excel avec un style amélioré
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Véhicules', index=False)
                    
                    # Récupérer la feuille de calcul
                    worksheet = writer.sheets['Véhicules']
                    
                    # Ajuster la largeur des colonnes
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(col)
                        )
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

                    # Styliser l'en-tête
                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='87CEFA',  # Bleu ciel
                            end_color='87CEFA',
                            fill_type='solid'
                        )

                messagebox.showinfo(
                    "Succès",
                    f"Les données ont été exportées avec succès vers:\n{file_path}"
                )
        except Exception as e:
            messagebox.showerror(
                "Erreur",
                f"Une erreur est survenue lors de l'exportation:\n{str(e)}"
            )

    def show_add_vehicle_form(self):
        # Créer une nouvelle fenêtre
        add_window = ctk.CTkToplevel(self)
        add_window.title("Ajouter un véhicule")
        add_window.geometry("400x550")

        # Variables pour stocker les entrées
        matricule_var = tk.StringVar()
        marque_var = tk.StringVar()
        modele_var = tk.StringVar()
        annee_var = tk.StringVar()
        etat_var = tk.StringVar(value="Disponible")
        kilometrage_var = tk.StringVar()

        # Création du formulaire
        ctk.CTkLabel(add_window, text="Matricule:").pack(pady=5)
        matricule_entry = ctk.CTkEntry(add_window, textvariable=matricule_var)
        matricule_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Marque:").pack(pady=5)
        marque_entry = ctk.CTkEntry(add_window, textvariable=marque_var)
        marque_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Modèle:").pack(pady=5)
        modele_entry = ctk.CTkEntry(add_window, textvariable=modele_var)
        modele_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Année:").pack(pady=5)
        annee_entry = ctk.CTkEntry(add_window, textvariable=annee_var)
        annee_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="État:").pack(pady=5)
        etat_combo = ttk.Combobox(add_window, textvariable=etat_var,
                                 values=["Disponible", "En mission", "En maintenance"])
        etat_combo.pack(pady=5)

        ctk.CTkLabel(add_window, text="Kilométrage:").pack(pady=5)
        kilometrage_entry = ctk.CTkEntry(add_window, textvariable=kilometrage_var)
        kilometrage_entry.pack(pady=5)

        # Bouton de confirmation
        def confirm_add():
            # Vérification des champs
            if not all([matricule_var.get(), marque_var.get(), modele_var.get(), 
                       annee_var.get(), etat_var.get(), kilometrage_var.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            # Ajouter le véhicule
            self.vehicle_tree.insert('', 'end', values=(
                matricule_var.get(),
                marque_var.get(),
                modele_var.get(),
                annee_var.get(),
                etat_var.get(),
                kilometrage_var.get(),
                datetime.now().strftime("%Y-%m-%d")
            ))
            add_window.destroy()
            messagebox.showinfo("Succès", "Véhicule ajouté avec succès")

        ctk.CTkButton(add_window, text="Confirmer", command=confirm_add).pack(pady=20)

    def show_modify_vehicle_form(self):
        selected = self.vehicle_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un véhicule à modifier")
            return

        # Récupérer les données du véhicule sélectionné
        vehicle_data = self.vehicle_tree.item(selected[0])['values']

        # Créer une nouvelle fenêtre
        modify_window = ctk.CTkToplevel(self)
        modify_window.title("Modifier un véhicule")
        modify_window.geometry("400x550")

        # Variables pour stocker les entrées
        matricule_var = tk.StringVar(value=vehicle_data[0])
        marque_var = tk.StringVar(value=vehicle_data[1])
        modele_var = tk.StringVar(value=vehicle_data[2])
        annee_var = tk.StringVar(value=vehicle_data[3])
        etat_var = tk.StringVar(value=vehicle_data[4])
        kilometrage_var = tk.StringVar(value=vehicle_data[5])

        # Création du formulaire
        ctk.CTkLabel(modify_window, text="Matricule:").pack(pady=5)
        matricule_entry = ctk.CTkEntry(modify_window, textvariable=matricule_var)
        matricule_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Marque:").pack(pady=5)
        marque_entry = ctk.CTkEntry(modify_window, textvariable=marque_var)
        marque_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Modèle:").pack(pady=5)
        modele_entry = ctk.CTkEntry(modify_window, textvariable=modele_var)
        modele_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Année:").pack(pady=5)
        annee_entry = ctk.CTkEntry(modify_window, textvariable=annee_var)
        annee_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="État:").pack(pady=5)
        etat_combo = ttk.Combobox(modify_window, textvariable=etat_var,
                                 values=["Disponible", "En mission", "En maintenance"])
        etat_combo.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Kilométrage:").pack(pady=5)
        kilometrage_entry = ctk.CTkEntry(modify_window, textvariable=kilometrage_var)
        kilometrage_entry.pack(pady=5)

        # Bouton de confirmation
        def confirm_modify():
            # Vérification des champs
            if not all([matricule_var.get(), marque_var.get(), modele_var.get(), 
                       annee_var.get(), etat_var.get(), kilometrage_var.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            # Modifier le véhicule
            self.vehicle_tree.item(selected[0], values=(
                matricule_var.get(),
                marque_var.get(),
                modele_var.get(),
                annee_var.get(),
                etat_var.get(),
                kilometrage_var.get(),
                datetime.now().strftime("%Y-%m-%d")
            ))
            modify_window.destroy()
            messagebox.showinfo("Succès", "Véhicule modifié avec succès")

        ctk.CTkButton(modify_window, text="Confirmer", command=confirm_modify).pack(pady=20)

    def delete_vehicle(self):
        selected = self.vehicle_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un véhicule à supprimer")
            return

        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir supprimer ce véhicule ?"):
            self.vehicle_tree.delete(selected[0])
            messagebox.showinfo("Succès", "Véhicule supprimé avec succès")



    def load_vehicle_data(self):
        # Nettoyer les données existantes
        for item in self.vehicle_tree.get_children():
            self.vehicle_tree.delete(item)

        # Données d'exemple
        sample_data = [
            ("AB-123-CD", "Renault", "Master", "2020", "Disponible", "50000", "2024-01-15"),
            ("EF-456-GH", "Peugeot", "Boxer", "2021", "En mission", "35000", "2024-02-20"),
            ("IJ-789-KL", "Mercedes", "Sprinter", "2019", "En maintenance", "75000", "2024-03-01")
        ]

        for vehicle in sample_data:
            self.vehicle_tree.insert('', 'end', values=vehicle)

    def employee(self):
        self.clear_content()
        self.emp_search_var = tk.StringVar(master=self)
        self.emp_role_var = tk.StringVar(master=self, value="Tous")
        
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Employés",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche
        search_label = ctk.CTkLabel(search_frame, text="Rechercher:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        self.emp_search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
         
            width=200,
            placeholder_text="Nom ou ID..."
        )
        search_entry.pack(side="left", padx=5)
        role_label = ctk.CTkLabel(search_frame, text="Poste:", font=("Arial", 12))
        role_label.pack(side="left", padx=5)
        
        self.emp_role_var = tk.StringVar(value="Tous")
        role_combo = ttk.Combobox(
            search_frame,
            
            values=["Tous", "Livreur", "Administrateur", "Manager", "Support client"],
            width=15,
            state="readonly"
        )
        role_combo.pack(side="left", padx=5)
        def search_employee():
           search_term = search_entry.get().lower()
           role_filter = role_combo .get()

           for item in self.emp_tree.get_children():
            values = self.emp_tree.item(item)['values']
            show = True

            if search_term:
                show = False
                for value in values[:3]:  # Recherche dans ID, Nom et Prénom
                    if str(value).lower().find(search_term) != -1:
                        show = True
                        break
            
            if role_filter != "Tous" and values[5] != role_filter:
                show = False

            if show:
                self.emp_tree.reattach(item, '', 'end')
            else:
                self.emp_tree.detach(item)

        # Filtre par poste
        

        # Boutons de recherche et réinitialisation
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_employee,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)

        reset_button = ctk.CTkButton(
            search_frame,
            text="↺ Réinitialiser",
            command=self.reset_employee_search,
            width=100,
            fg_color=self.orange
        )
        reset_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des employés
        columns = ("ID", "Nom", "Prénom", "Email", "Téléphone", "Poste", "Date d'embauche", "Statut")
        self.emp_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.emp_tree.heading(col, text=col, command=lambda c=col: self.sort_employees(c))
            width = 150 if col in ["Nom", "Prénom", "Email"] else 100
            self.emp_tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.emp_tree.yview)
        self.emp_tree.configure(yscrollcommand=scrollbar.set)
        
        self.emp_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame pour les boutons
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons d'action
        add_button = ctk.CTkButton(
            button_frame,
            text="➕ Ajouter",
            command=self.show_add_employee_form,
            fg_color="green",
            width=120,
            height=32
        )
        add_button.pack(side="left", padx=5)

        modify_button = ctk.CTkButton(
            button_frame,
            text="✏️ Modifier",
            command=self.show_modify_employee_form,
            fg_color="orange",
            width=120,
            height=32
        )
        modify_button.pack(side="left", padx=5)

        delete_button = ctk.CTkButton(
            button_frame,
            text="🗑️ Supprimer",
            command=self.delete_employee,
            fg_color="red",
            width=120,
            height=32
        )
        delete_button.pack(side="left", padx=5)

        export_button = ctk.CTkButton(
            button_frame,
            text="📊 Exporter Excel",
            command=self.export_to_excel_emp,
            fg_color="#2E7D32",
            width=120,
            height=32
        )
        export_button.pack(side="left", padx=5)

        # Charger les données
        self.load_employee_data()

    def show_add_employee_form(self):
        add_window = ctk.CTkToplevel(self)
        add_window.title("Ajouter un employé")
        add_window.geometry("400x600")

        # Variables
        id_var = tk.StringVar()
        nom_var = tk.StringVar()
        prenom_var = tk.StringVar()
        email_var = tk.StringVar()
        tel_var = tk.StringVar()
        poste_var = tk.StringVar(value="Livreur")
        statut_var = tk.StringVar(value="Actif")

        # Formulaire
        ctk.CTkLabel(add_window, text="ID:").pack(pady=5)
        id_entry = ctk.CTkEntry(add_window, textvariable=id_var)
        id_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Nom:").pack(pady=5)
        nom_entry = ctk.CTkEntry(add_window, textvariable=nom_var)
        nom_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Prénom:").pack(pady=5)
        prenom_entry = ctk.CTkEntry(add_window, textvariable=prenom_var)
        prenom_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Email:").pack(pady=5)
        email_entry = ctk.CTkEntry(add_window, textvariable=email_var)
        email_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Téléphone:").pack(pady=5)
        tel_entry = ctk.CTkEntry(add_window, textvariable=tel_var)
        tel_entry.pack(pady=5)

        ctk.CTkLabel(add_window, text="Poste:").pack(pady=5)
        poste_combo = ttk.Combobox(add_window, textvariable=poste_var,
                                  values=["Livreur", "Administrateur", "Manager", "Support client"])
        poste_combo.pack(pady=5)

        ctk.CTkLabel(add_window, text="Statut:").pack(pady=5)
        statut_combo = ttk.Combobox(add_window, textvariable=statut_var,
                                   values=["Actif", "Inactif", "En congé"])
        statut_combo.pack(pady=5)

        def confirm_add():
            if not all([id_var.get(), nom_var.get(), prenom_var.get(), email_var.get(), 
                       tel_var.get(), poste_var.get(), statut_var.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            self.emp_tree.insert('', 'end', values=(
                id_var.get(),
                nom_var.get(),
                prenom_var.get(),
                email_var.get(),
                tel_var.get(),
                poste_var.get(),
                datetime.now().strftime("%Y-%m-%d"),
                statut_var.get()
            ))
            add_window.destroy()
            messagebox.showinfo("Succès", "Employé ajouté avec succès")

        ctk.CTkButton(add_window, text="Confirmer", command=confirm_add).pack(pady=20)

    def show_modify_employee_form(self):
        selected = self.emp_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un employé à modifier")
            return

        emp_data = self.emp_tree.item(selected[0])['values']
        
        modify_window = ctk.CTkToplevel(self)
        modify_window.title("Modifier un employé")
        modify_window.geometry("400x600")

        # Variables
        id_var = tk.StringVar(value=emp_data[0])
        nom_var = tk.StringVar(value=emp_data[1])
        prenom_var = tk.StringVar(value=emp_data[2])
        email_var = tk.StringVar(value=emp_data[3])
        tel_var = tk.StringVar(value=emp_data[4])
        poste_var = tk.StringVar(value=emp_data[5])
        statut_var = tk.StringVar(value=emp_data[7])

        # Formulaire
        ctk.CTkLabel(modify_window, text="ID:").pack(pady=5)
        id_entry = ctk.CTkEntry(modify_window, textvariable=id_var, state="disabled")
        id_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Nom:").pack(pady=5)
        nom_entry = ctk.CTkEntry(modify_window, textvariable=nom_var)
        nom_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Prénom:").pack(pady=5)
        prenom_entry = ctk.CTkEntry(modify_window, textvariable=prenom_var)
        prenom_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Email:").pack(pady=5)
        email_entry = ctk.CTkEntry(modify_window, textvariable=email_var)
        email_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Téléphone:").pack(pady=5)
        tel_entry = ctk.CTkEntry(modify_window, textvariable=tel_var)
        tel_entry.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Poste:").pack(pady=5)
        poste_combo = ttk.Combobox(modify_window, textvariable=poste_var,
                                  values=["Livreur", "Administrateur", "Manager", "Support client"])
        poste_combo.pack(pady=5)

        ctk.CTkLabel(modify_window, text="Statut:").pack(pady=5)
        statut_combo = ttk.Combobox(modify_window, textvariable=statut_var,
                                   values=["Actif", "Inactif", "En congé"])
        statut_combo.pack(pady=5)

        def confirm_modify():
            if not all([nom_var.get(), prenom_var.get(), email_var.get(), 
                       tel_var.get(), poste_var.get(), statut_var.get()]):
                messagebox.showwarning("Erreur", "Veuillez remplir tous les champs")
                return

            self.emp_tree.item(selected[0], values=(
                id_var.get(),
                nom_var.get(),
                prenom_var.get(),
                email_var.get(),
                tel_var.get(),
                poste_var.get(),
                emp_data[6],  # Garder la date d'embauche originale
                statut_var.get()
            ))
            modify_window.destroy()
            messagebox.showinfo("Succès", "Employé modifié avec succès")

        ctk.CTkButton(modify_window, text="Confirmer", command=confirm_modify).pack(pady=20)

    def delete_employee(self):
        selected = self.emp_tree.selection()
        if not selected:
            messagebox.showwarning("Erreur", "Veuillez sélectionner un employé à supprimer")
            return

        emp_data = self.emp_tree.item(selected[0])['values']
        if messagebox.askyesno("Confirmation", f"Êtes-vous sûr de vouloir supprimer l'employé {emp_data[1]} {emp_data[2]} ?"):
            self.emp_tree.delete(selected[0])
            messagebox.showinfo("Succès", "Employé supprimé avec succès")

    

    def reset_employee_search(self):
        self.emp_search_var.set("")
        self.emp_role_var.set("Tous")
        self.load_employee_dataem()

    def load_employee_data(self):
        for item in self.emp_tree.get_children():
            self.emp_tree.delete(item)

        sample_data = [
            ("EMP001", "Dupont", "Jean", "jean.dupont@email.com", "0123456789", "Livreur", "2023-01-15", "Actif"),
            ("EMP002", "Martin", "Sophie", "sophie.martin@email.com", "0234567890", "Manager", "2022-06-20", "Actif"),
            ("EMP003", "Bernard", "Pierre", "pierre.bernard@email.com", "0345678901", "Support client", "2023-03-10", "En congé")
        ]

        for emp in sample_data:
            self.emp_tree.insert('', 'end', values=emp)

    def export_to_excel_emp(self):
        try:
            data = []
            columns = ["ID", "Nom", "Prénom", "Email", "Téléphone", "Poste", "Date d'embauche", "Statut"]
            
            for item in self.emp_tree.get_children():
                values = self.emp_tree.item(item)['values']
                data.append(values)

            df = pd.DataFrame(data, columns=columns)

            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Enregistrer le fichier Excel"
            )

            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Employés', index=False)
                    
                    worksheet = writer.sheets['Employés']
                    
                    for idx, col in enumerate(df.columns):
                        max_length = max(df[col].astype(str).apply(len).max(), len(col))
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='87CEFA',
                            end_color='87CEFA',
                            fill_type='solid'
                        )

                messagebox.showinfo("Succès", f"Les données ont été exportées avec succès vers:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation:\n{str(e)}")

    def sort_employees(self, column):
        """Trier le tableau par colonne"""
        l = [(self.emp_tree.set(k, column), k) for k in self.emp_tree.get_children('')]
        l.sort(reverse=getattr(self, '_sort_reverse', False))
        self._sort_reverse = not getattr(self, '_sort_reverse', False)
        
        for index, (_, k) in enumerate(l):
            self.emp_tree.move(k, '', index)
    def manage_clients(self):
       
        self.clear_content()
       
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Gestion des Clients",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Barre de recherche
        search_label = ctk.CTkLabel(search_frame, text="Rechercher:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
       
        search_entry = ctk.CTkEntry(
            search_frame,
            width=200,
            placeholder_text="Nom ou ID..."
        )
        search_entry.pack(side="left", padx=5)
        type_label = ctk.CTkLabel(search_frame, text="Type:", font=("Arial", 12))
        type_label.pack(side="left", padx=5)
        
        
        type_combo = ttk.Combobox(
            search_frame,
            values=["Tous", "Particulier", "Professionnel", "VIP"],
            width=15,
            state="readonly"
        )
        type_combo.pack(side="left", padx=5)

        def search_client():
           search_term = search_entry.get().lower()
           type_filter = type_combo.get()

           for item in self.client_tree.get_children():
            values = self.client_tree.item(item)['values']
            show = True

            if search_term:
                show = False
                for value in values[:3]:  # Recherche dans ID, Nom et Prénom
                    if str(value).lower().find(search_term) != -1:
                        show = True
                        break
            
            if type_filter != "Tous" and values[5] != type_filter:
                show = False

            if show:
                self.client_tree.reattach(item, '', 'end')
            else:
                self.client_tree.detach(item)

        # Boutons de recherche et réinitialisation
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_client,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)
        def reset_client_search():
            search_entry.insert("0","")
            type_combo.set("Tous")
            self.load_client_data()

        reset_button = ctk.CTkButton(
            search_frame,
            text="↺ Réinitialiser",
            command=reset_client_search,
            width=100,
            fg_color=self.orange
        )
        reset_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des clients
        columns = ("ID", "Nom", "Prénom", "Email", "Téléphone", "Type", "Date d'inscription", "Statut")
        self.client_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.client_tree.heading(col, text=col, command=lambda c=col: self.sort_clients(c))
            width = 150 if col in ["Nom", "Prénom", "Email"] else 100
            self.client_tree.column(col, width=width, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.client_tree.yview)
        self.client_tree.configure(yscrollcommand=scrollbar.set)
        
        self.client_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame pour les boutons
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        

        

        export_button = ctk.CTkButton(
            button_frame,
            text="📊 Exporter Excel",
            command=self.export_to_excel_client,
            fg_color="#2E7D32",
            width=120,
            height=32
        )
        export_button.pack(side="left", padx=5)

        # Charger les données
        self.load_client_data()
        

        # Trait de séparation gris
    def check_tables(self):
     import sqlite3
     try:
        conn = sqlite3.connect("G:\\.shortcut-targets-by-id\\1EnGOIt2FrSKgqS9VvWKycgALlkupt7U_\\projet final\\suivi_coli.db")
        cursor = conn.cursor()
        
        # Vérifier la table user
        cursor.execute("SELECT COUNT(*) FROM user")
        user_count = cursor.fetchone()[0]
        print(f"Nombre d'utilisateurs: {user_count}")
        
        # Vérifier la table individu
        cursor.execute("SELECT COUNT(*) FROM individu")
        individu_count = cursor.fetchone()[0]
        print(f"Nombre d'individus: {individu_count}")
        
        # Vérifier la table société
        cursor.execute("SELECT COUNT(*) FROM societé")
        societe_count = cursor.fetchone()[0]
        print(f"Nombre de sociétés: {societe_count}")
        
     except sqlite3.Error as e:
        print(f"Erreur lors de la vérification des tables: {e}")
     finally:
        if 'conn' in locals():
            conn.close()
    
    def load_client_data(self):
     import sqlite3
     # Nettoyer le tableau existant
     for item in self.client_tree.get_children():
        self.client_tree.delete(item)

     try:
        # Connexion à la base de données
        conn = sqlite3.connect("G:\\.shortcut-targets-by-id\\1EnGOIt2FrSKgqS9VvWKycgALlkupt7U_\\projet final\\suivi_coli.db")
        cursor = conn.cursor()

        # Requête pour les individus
        cursor.execute("""
            SELECT u.id_user, 
                   i.nm_individu, 
                   i.pr_individu, 
                   i.email_individu, 
                   i.numTel_individu, 
                   'Particulier' as type_client,
                   i.ville,
                   'Actif' as statut
            FROM user u
            JOIN individu i ON u.cin_individu = i.cin_individu
            WHERE u.cin_individu IS NOT NULL
        """)
        individus = cursor.fetchall()

        # Requête pour les sociétés
        cursor.execute("""
            SELECT u.id_user,
                   s.nom_societ,
                   '' as prenom,
                   s.email_societ,
                   s.num_telsoc,
                   'Professionnel' as type_client,
                   s.ville,
                   'Actif' as statut
            FROM user u
            JOIN societé s ON u.ice_societ = s.ice_societ
            WHERE u.ice_societ IS NOT NULL
        """)
        societes = cursor.fetchall()

        # Combiner et afficher les résultats
        all_clients = individus + societes
        if all_clients:
         for client in all_clients:
            self.client_tree.insert('', 'end', values=client)

        # Si aucun client trouvé
        else:
            self.client_tree.insert('', 'end', values=('Aucun client trouvé', '', '', '', '', '', '', ''))

     except sqlite3.Error as e:
        messagebox.showerror("Erreur", f"Erreur lors du chargement des clients: {e}")
        self.client_tree.insert('', 'end', values=('Erreur de chargement', '', '', '', '', '', '', ''))
    
     finally:
        if 'conn' in locals():
            conn.close()

    def export_to_excel_client(self):
        try:
            data = []
            columns = ["ID", "Nom", "Prénom", "Email", "Téléphone", "Type", "Date d'inscription", "Statut"]
            
            for item in self.client_tree.get_children():
                values = self.client_tree.item(item)['values']
                data.append(values)

            df = pd.DataFrame(data, columns=columns)

            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Enregistrer le fichier Excel"
            )

            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Clients', index=False)
                    
                    worksheet = writer.sheets['Clients']
                    
                    for idx, col in enumerate(df.columns):
                        max_length = max(df[col].astype(str).apply(len).max(), len(col))
                        worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='87CEFA',
                            end_color='87CEFA',
                            fill_type='solid'
                        )

                messagebox.showinfo("Succès", f"Les données ont été exportées avec succès vers:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation:\n{str(e)}")

    def sort_clients(self, column):
        """Trier le tableau par colonne"""
        l = [(self.client_tree.set(k, column), k) for k in self.client_tree.get_children('')]
        l.sort(reverse=getattr(self, '_sort_reverse', False))
        self._sort_reverse = not getattr(self, '_sort_reverse', False)
        
        for index, (_, k) in enumerate(l):
            self.client_tree.move(k, '', index)
    


    def tracking_packages(self):
        self.clear_content()
        
        # Frame titre
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=60)
        title_frame.pack(fill="x", pady=(0, 20))
        title_label = ctk.CTkLabel(
            title_frame,
            text="Suivi des Colis",
            font=("Arial", 24, "bold"),
            text_color="white"
        )
        title_label.pack(pady=10)

        # Frame de recherche
        search_frame = ctk.CTkFrame(self.content_frame)
        search_frame.pack(fill="x", padx=10, pady=5)

        # Recherche par ID
        search_label = ctk.CTkLabel(search_frame, text="Rechercher ID:", font=("Arial", 12))
        search_label.pack(side="left", padx=5)
        
        self.search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=self.search_var,
            width=200,
            placeholder_text="Entrez l'ID..."
        )
        search_entry.pack(side="left", padx=5)
        def search_package():
            """Rechercher une demande par ID"""
            search_term = search_entry.get().upper()
            print(search_term)
        
            if not search_term:
                messagebox.showwarning("Attention", "Veuillez entrer un ID")
                return
            
        
        # Convertir l'ID recherché en minuscules pour une recherche insensible à la casse
        
            found = False
        
            for item in self.tracking_tree.get_children():
                values = self.tracking_tree.item(item)['values']
                # Convertir l'ID de la ligne en minuscules et en string pour la comparaison
                current_id = str(values[0]).upper()
                print(current_id)
            
                if current_id == search_term:
                # Sélectionner et mettre en évidence l'élément trouvé
                    self.tracking_tree.selection_set(item)
                    self.tracking_tree.see(item)
                    found = True
                    break
        
            if not found:
                messagebox.showwarning("Recherche", "Aucune demande trouvée avec cet ID")



        # Bouton de recherche
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=search_package,
            width=100,
            fg_color=self.orange
        )
        search_button.pack(side="left", padx=5)

        # Frame pour le tableau
        table_frame = ctk.CTkFrame(self.content_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des colis
        columns = ("ID", "Client", "Ville Départ", "Ville Arrivée", "État", "Latitude", "Longitude", "Date MAJ")
        self.tracking_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.tracking_tree.heading(col, text=col)
            self.tracking_tree.column(col, width=120, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tracking_tree.yview)
        self.tracking_tree.configure(yscrollcommand=scrollbar.set)
        
        self.tracking_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Frame pour les boutons d'action
        button_frame = ctk.CTkFrame(self.content_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Boutons d'action
        update_city_btn = ctk.CTkButton(
            button_frame,
            text="🌍 Modifier Ville",
            command=self.update_city_location,
            fg_color=self.orange,
            width=150
        )
        update_city_btn.pack(side="left", padx=5)

        update_status_btn = ctk.CTkButton(
            button_frame,
            text="📝 Modifier État",
            command=self.update_status,
            fg_color=self.blue_ciel,
            width=150
        )
        update_status_btn.pack(side="left", padx=5)

        # Charger les données exemple
        self.load_sample_datatr()

        
    def load_sample_datatr(self):
        # Nettoyer les données existantes
        for item in self.tracking_tree.get_children():
            self.tracking_tree.delete(item)
            
        # Données exemple
        data = [
            ("COL001", "Ahmed Alami", "Casablanca", "Rabat", "En transit", 33.5731, -7.5898, "2024-03-15 10:00"),
            ("COL002", "Sara Bennani", "Marrakech", "Tanger", "En livraison", 31.6295, -7.9811, "2024-03-15 11:30"),
            ("COL003", "Karim Idrissi", "Fès", "Agadir", "En attente", 34.0333, -5.0000, "2024-03-15 09:15")
        ]

        # Insérer les nouvelles données
        for row in data:
            self.tracking_tree.insert('', 'end', values=row)

    def update_city_location(self):
        selected = self.tracking_tree.selection()
        if not selected:
            messagebox.showwarning("Attention", "Veuillez sélectionner un colis")
            return

        # Fenêtre de modification
        update_window = ctk.CTkToplevel()
        update_window.title("Modifier la ville")
        update_window.geometry("300x200")

        # Liste des villes avec leurs coordonnées
        villes = {
            "Casablanca": (33.5731, -7.5898),
            "Rabat": (34.0209, -6.8416),
            "Marrakech": (31.6295, -7.9811),
            "Fès": (34.0333, -5.0000),
            "Tanger": (35.7595, -5.8340),
            "Agadir": (30.4278, -9.5981)
        }

        ville_var = tk.StringVar()
        ville_combo = ttk.Combobox(
            update_window,
            textvariable=ville_var,
            values=list(villes.keys()),
            state="readonly"
        )
        ville_combo.pack(pady=20)

        def confirm_update():
            if ville_var.get():
                lat, lng = villes[ville_var.get()]
                item = selected[0]
                values = list(self.tracking_tree.item(item)['values'])
                values[2] = ville_var.get()  # Ville
                values[5] = lat  # Latitude
                values[6] = lng  # Longitude
                values[7] = datetime.now().strftime("%Y-%m-%d %H:%M")  # Date MAJ
                self.tracking_tree.item(item, values=values)
                update_window.destroy()

        ctk.CTkButton(
            update_window,
            text="Confirmer",
            command=confirm_update
        ).pack(pady=10)

    def update_status(self):
        selected = self.tracking_tree.selection()
        if not selected:
            messagebox.showwarning("Attention", "Veuillez sélectionner un colis")
            return

        # Fenêtre de modification
        status_window = ctk.CTkToplevel()
        status_window.title("Modifier l'état")
        status_window.geometry("300x200")

        statuts = ["En attente", "En transit", "En livraison", "Livré", "Retourné"]
        status_var = tk.StringVar()
        status_combo = ttk.Combobox(
            status_window,
            textvariable=status_var,
            values=statuts,
            state="readonly"
        )
        status_combo.pack(pady=20)

        def confirm_status():
            if status_var.get():
                item = selected[0]
                values = list(self.tracking_tree.item(item)['values'])
                values[4] = status_var.get()  # État
                values[7] = datetime.now().strftime("%Y-%m-%d %H:%M")  # Date MAJ
                self.tracking_tree.item(item, values=values)
                status_window.destroy()

        ctk.CTkButton(
            status_window,
            text="Confirmer",
            command=confirm_status
        ).pack(pady=10)
    
            
    

    def sort_column(self, col):
        l = [(self.tracking_tree.set(k, col), k) for k in self.tracking_tree.get_children('')]
        l.sort(reverse=getattr(self, f'_sort_reverse_{col}', False))
        setattr(self, f'_sort_reverse_{col}', not getattr(self, f'_sort_reverse_{col}', False))
        
        for index, (_, k) in enumerate(l):
            self.tracking_tree.move(k, '', index)

    def load_sample_data(self):
        # Nettoyer les données existantes
        for item in self.tracking_tree.get_children():
            self.tracking_tree.delete(item)
            
        # Données exemple
        data = [
            ("COL001", "Ahmed Alami", "Casablanca", "Rabat", "En transit", 33.5731, -7.5898, "2024-03-15 10:00"),
            ("COL002", "Sara Bennani", "Marrakech", "Tanger", "En livraison", 31.6295, -7.9811, "2024-03-15 11:30"),
            ("COL003", "Karim Idrissi", "Fès", "Agadir", "En attente", 34.0333, -5.0000, "2024-03-15 09:15")
        ]

        # Insérer les nouvelles données
        for row in data:
            self.tracking_tree.insert('', 'end', values=row)
    def view_statistics(self):
        self.clear_content()
        
        # Frame titre avec style moderne
        title_frame = ctk.CTkFrame(self.content_frame, fg_color=self.blue_ciel, height=70)
        title_frame.pack(fill="x", pady=(0, 20))
        
        # Titre avec icône
        title_label = ctk.CTkLabel(
            title_frame,
            text="📊 Tableau de Bord",
            font=("Helvetica", 28, "bold"),
            text_color="white"
        )
        title_label.pack(pady=15)

        # Frame pour les filtres avec style moderne
        filter_frame = ctk.CTkFrame(self.content_frame, fg_color="#f0f0f0", corner_radius=15)
        filter_frame.pack(fill="x", padx=20, pady=10)

        # Container pour les filtres
        filter_container = ctk.CTkFrame(filter_frame, fg_color="transparent")
        filter_container.pack(pady=10, padx=20)

        # Sélection de la période avec style
        period_label = ctk.CTkLabel(
            filter_container, 
            text="📅 Période:", 
            font=("Helvetica", 14, "bold"),
            text_color="#333"
        )
        period_label.pack(side="left", padx=(0,10))
        
        self.period_var = tk.StringVar(value="Ce mois")
        period_combo = ttk.Combobox(
            filter_container,
            textvariable=self.period_var,
            values=["Aujourd'hui", "Cette semaine", "Ce mois", "Cette année", "Personnalisé"],
            width=15,
            state="readonly"
        )
        period_combo.pack(side="left", padx=5)

        # Dates personnalisées
        date_frame = ctk.CTkFrame(filter_container, fg_color="transparent")
        date_frame.pack(side="left", padx=20)

        ctk.CTkLabel(
            date_frame, 
            text="Du:", 
            font=("Helvetica", 12),
            text_color="#333"
        ).pack(side="left", padx=5)
        
        self.date_debut = DateEntry(
            date_frame, 
            width=12, 
            background=self.blue_ciel,
            foreground='white',
            borderwidth=0
        )
        self.date_debut.pack(side="left", padx=5)
        
        ctk.CTkLabel(
            date_frame, 
            text="Au:", 
            font=("Helvetica", 12),
            text_color="#333"
        ).pack(side="left", padx=5)
        
        self.date_fin = DateEntry(
            date_frame, 
            width=12, 
            background=self.blue_ciel,
            foreground='white',
            borderwidth=0
        )
        self.date_fin.pack(side="left", padx=5)

        # Bouton actualiser et exporter
        button_frame = ctk.CTkFrame(filter_container, fg_color="transparent")
        button_frame.pack(side="left", padx=20)

        refresh_button = ctk.CTkButton(
            button_frame,
            text="🔄 Actualiser",
            command=self.update_statistics,
            font=("Helvetica", 12, "bold"),
            fg_color=self.blue_ciel,
            hover_color="#1976D2",
            width=120
        )
        refresh_button.pack(side="left", padx=5)

        export_button = ctk.CTkButton(
            button_frame,
            text="📥 Exporter",
            command=self.export_statistics,
            font=("Helvetica", 12, "bold"),
            fg_color=self.orange,
            hover_color="#F57C00",
            width=120
        )
        export_button.pack(side="left", padx=5)

        # Container principal
        main_container = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=10)

        # KPI Cards
        self.create_kpi_section(main_container)

        # Graphiques principaux
        
        
        charts_container = ctk.CTkFrame(main_container, fg_color="transparent")
        charts_container.pack(fill="both", expand=True, pady=10)

        # Créer les graphiques l'un après l'autre verticalement
        self.create_delivery_trend(charts_container)
        
        # Ajouter une légende pour le graphique d'évolution
        legend_frame1 = ctk.CTkFrame(charts_container, fg_color="white", corner_radius=10)
        legend_frame1.pack(fill="x", pady=(0, 20), padx=10)
        ctk.CTkLabel(
            legend_frame1,
            text="📈 Évolution des livraisons : La courbe bleue représente le nombre de colis livrés par jour\n"
                 "La zone ombrée indique la tendance générale",
            font=("Helvetica", 12),
            text_color="#666"
        ).pack(pady=10)

        self.create_city_analysis(charts_container)
        
        # Légende pour l'analyse par ville
        legend_frame2 = ctk.CTkFrame(charts_container, fg_color="white", corner_radius=10)
        legend_frame2.pack(fill="x", pady=(0, 20), padx=10)
        ctk.CTkLabel(
            legend_frame2,
            text="🏢 Distribution par ville : Les barres représentent le volume de livraisons par ville\n"
                 "Dégradé de bleu : Plus foncé = volume plus important",
            font=("Helvetica", 12),
            text_color="#666"
        ).pack(pady=10)

        self.create_status_distribution(charts_container)
        
        # Légende pour la distribution des états
        legend_frame3 = ctk.CTkFrame(charts_container, fg_color="white", corner_radius=10)
        legend_frame3.pack(fill="x", pady=(0, 20), padx=10)
        ctk.CTkLabel(
            legend_frame3,
            text="🔄 États des colis : \n"
                 "Bleu = En transit | Vert = Livré | Orange = En attente | Rouge = Retourné",
            font=("Helvetica", 12),
            text_color="#666"
        ).pack(pady=10)

        

    def create_kpi_section(self, parent):
        kpi_frame = ctk.CTkFrame(parent, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=10)

        kpi_data = [
            {
                "title": "📦 Total Colis",
                "value": "1,234",
                "change": "+12%",
                "color": self.blue_ciel
            },
            {
                "title": "✅ Taux de Livraison",
                "value": "95%",
                "change": "+3%",
                "color": self.green
            },
            {
                "title": "⏱️ Délai Moyen",
                "value": "2.3 jours",
                "change": "-0.5j",
                "color": self.orange
            },
            {
                "title": "❌ Retours",
                "value": "3%",
                "change": "-1%",
                "color": self.red
            }
        ]

        for kpi in kpi_data:
            self.create_kpi_card(kpi_frame, kpi)

    def create_kpi_card(self, parent, data):
        card = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=10,
            border_width=2,
            border_color="#e0e0e0"
        )
        card.pack(side="left", padx=10, fill="both", expand=True)

        title = ctk.CTkLabel(
            card,
            text=data["title"],
            font=("Helvetica", 14, "bold"),
            text_color=data["color"]
        )
        title.pack(pady=(15,5))

        value = ctk.CTkLabel(
            card,
            text=data["value"],
            font=("Helvetica", 24, "bold"),
            text_color="#333"
        )
        value.pack()

        change = ctk.CTkLabel(
            card,
            text=data["change"],
            font=("Helvetica", 12),
            text_color=self.green if "+" in data["change"] else self.red
        )
        change.pack(pady=(0,15))

    def create_delivery_trend(self, parent):
        chart_frame = self.create_chart_container(parent, "Évolution des Livraisons")
        
        # Données exemple
        dates = pd.date_range(start='2024-01-01', end='2024-03-15', freq='D')
        deliveries = np.random.randint(20, 50, size=len(dates))

        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(dates, deliveries, color=self.blue_ciel, linewidth=2)
        ax.fill_between(dates, deliveries, color=self.blue_ciel, alpha=0.2)
        
        ax.set_facecolor('white')
        fig.patch.set_facecolor('white')
        ax.grid(True, linestyle='--', alpha=0.3)
        plt.xticks(rotation=45)
        
        canvas = FigureCanvasTkAgg(fig, chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(pady=10, padx=10, fill="both", expand=True)

    def create_city_analysis(self, parent):
        chart_frame = self.create_chart_container(parent, "Livraisons par Ville")
        
        cities = ['Casablanca', 'Rabat', 'Marrakech', 'Tanger', 'Fès']
        deliveries = [150, 100, 80, 60, 40]

        fig, ax = plt.subplots(figsize=(8, 4))
        bars = ax.bar(cities, deliveries)
        
        ax.set_facecolor('white')
        fig.patch.set_facecolor('white')
        
        colors = plt.cm.Blues(np.linspace(0.4, 0.8, len(cities)))
        for bar, color in zip(bars, colors):
            bar.set_color(color)

        plt.xticks(rotation=45)
        ax.grid(True, linestyle='--', alpha=0.3)
        
        canvas = FigureCanvasTkAgg(fig, chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(pady=10, padx=10, fill="both", expand=True)

    def create_status_distribution(self, parent):
        chart_frame = self.create_chart_container(parent, "Distribution des États")
        
        status_data = {
            'En transit': 30,
            'Livré': 45,
            'En attente': 15,
            'Retourné': 10
        }
        
        fig, ax = plt.subplots(figsize=(8, 4))
        colors = [self.blue_ciel, self.green, self.orange, self.red]
        
        wedges, texts, autotexts = ax.pie(
            status_data.values(),
            labels=status_data.keys(),
            colors=colors,
            autopct='%1.1f%%',
            startangle=90
        )
        
        plt.setp(autotexts, size=8, weight="bold")
        plt.setp(texts, size=8)
        
        canvas = FigureCanvasTkAgg(fig, chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(pady=10, padx=10, fill="both", expand=True)

    def create_performance_metrics(self, parent):
        chart_frame = self.create_chart_container(parent, "Métriques de Performance")
        
        metrics = {
            'Délai < 24h': 85,
            'Satisfaction': 92,
            'Premier essai': 78,
            'Sans incident': 95
        }
        
        fig, ax = plt.subplots(figsize=(8, 4))
        y_pos = np.arange(len(metrics))
        
        bars = ax.barh(y_pos, list(metrics.values()))
        ax.set_yticks(y_pos)
        ax.set_yticklabels(metrics.keys())
        
        ax.set_facecolor('white')
        fig.patch.set_facecolor('white')
        
        # Colorer les barres selon la performance
        for i, bar in enumerate(bars):
            if metrics[list(metrics.keys())[i]] >= 90:
                bar.set_color(self.green)
            elif metrics[list(metrics.keys())[i]] >= 75:
                bar.set_color(self.blue_ciel)
            else:
                bar.set_color(self.orange)
        
        ax.grid(True, linestyle='--', alpha=0.3)
        
        # Ajouter les valeurs sur les barres
        for i, v in enumerate(metrics.values()):
            ax.text(v + 1, i, f'{v}%', va='center')
        
        canvas = FigureCanvasTkAgg(fig, chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(pady=10, padx=10, fill="both", expand=True)

    def create_chart_container(self, parent, title):
        container = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=10,
            border_width=2,
            border_color="#e0e0e0"
        )
        container.pack(fill="both", expand=True, pady=5)

        title_label = ctk.CTkLabel(
            container,
            text=title,
            font=("Helvetica", 16, "bold"),
            text_color="#333"
        )
        title_label.pack(pady=10)

        return container

    def update_statistics(self):
        # Mettre à jour les données selon la période sélectionnée
        period = self.period_var.get()
        start_date = self.date_debut.get_date()
        end_date = self.date_fin.get_date()
        
        # Actualiser chaque section
        self.clear_content()
        self.view_statistics()
        messagebox.showinfo("Mise à jour", "Statistiques actualisées avec succès!")

    def export_statistics(self):
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx"), ("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Exporter les statistiques"
            )
            
            if file_path:
                # Créer un rapport Excel
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Exporter les KPIs
                    kpi_data = pd.DataFrame({
                        "Métrique": ["Total Colis", "Taux de Livraison", "Délai Moyen", "Taux de Retour"],
                        "Valeur": ["1,234", "95%", "2.3 jours", "3%"],
                        "Évolution": ["+12%", "+3%", "-0.5j", "-1%"]
                    })
                    kpi_data.to_excel(writer, sheet_name='KPIs', index=False)
                    
                    # Exporter les autres données...
                    
                messagebox.showinfo("Succès", f"Données exportées vers:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'exportation:\n{str(e)}")

    def logout(self):
        # Déconnexion (simulé)
        messagebox.showinfo("Déconnexion", "Vous êtes déconnecté.")

    def clear_content(self):
        # Supprimer tous les widgets dans le frame de contenu
        for widget in self.content_frame.winfo_children():
            widget.destroy()

    def send_notifications(self):
        """Interface pour l'envoi et la gestion des notifications"""
        self.clear_content()
    
        # Initialisation des données d'exemple si elles n'existent pas
        if not hasattr(self, 'mock_notifications'):
            self.mock_notifications = [
                {
                    'id': 'NOT001',
                    'colis_id': 'COL001',
                    'type': 'Livraison',
                    'message': 'Votre colis sera livré aujourd\'hui',
                    'destinataire': 'Ahmed Alami',
                    'adresse': '123 Rue Hassan II, Casablanca',
                    'telephone': '0612345678',
                    'montant': 250.00,
                    'etat': 'En cours',
                    'date_creation': '2024-03-15 09:30',
                    'date_livraison_debut': '2024-03-16 09:00',
                    'date_livraison_fin': '2024-03-16 12:00',
                    'priorite': 'Normale',
                    'lu': False
                },
                {
                    'id': 'NOT002',
                    'colis_id': 'COL002',
                    'type': 'Retard',
                    'message': 'Retard de livraison prévu',
                    'destinataire': 'Sara Bennani',
                    'adresse': '45 Avenue FAR, Rabat',
                    'telephone': '0623456789',
                    'montant': 180.00,
                    'etat': 'En attente',
                    'date_creation': '2024-03-15 10:15',
                    'date_livraison_debut': '2024-03-16 14:00',
                    'date_livraison_fin': '2024-03-16 17:00',
                    'priorite': 'Haute',
                    'lu': False
                }
            ]

        # ... (reste du code inchangé jusqu'à la création du Treeview)
        
        # Frame principal avec fond blanc
        main_container = ctk.CTkFrame(self.content_frame, fg_color="white")
        main_container.pack(fill="both", expand=True)

        # Frame titre
        title_frame = ctk.CTkFrame(main_container, fg_color=self.blue_ciel, height=70)
        title_frame.pack(fill="x", pady=(0, 20))
        
        title_label = ctk.CTkLabel(
            title_frame,
            text="🔔 Gestion des Notifications",
            font=("Helvetica", 28, "bold"),
            text_color="white"
        )
        title_label.pack(pady=15)

        # Frame de recherche et filtres
        search_frame = ctk.CTkFrame(main_container, fg_color="white")
        search_frame.pack(fill="x", padx=20, pady=10)

        # Champ de recherche
        # Champ de recherche
        search_entry = ctk.CTkEntry(
            search_frame, 
            placeholder_text="Rechercher une notification...",
            width=300
            
        )
        search_entry.pack(side="left", padx=5)
        
        
        def rechercher_notifications():
           """Rechercher dans les notifications"""
           search_term = search_entry.get().lower().strip()
        
        # Si la recherche est vide, afficher toutes les notifications
           if not search_term:
               self.actualiser_notifications()
               return

        # Effacer le tableau
           for item in self.notif_tree.get_children():
               self.notif_tree.delete(item)

        # Filtrer et afficher les résultats
           found = False
           for notif in self.mock_notifications:
              # Recherche dans plusieurs champs
              searchable_fields = [
                str(notif['colis_id']).lower(),
                str(notif['destinataire']).lower(),
                str(notif['type']).lower(),
                str(notif['etat']).lower(),
                str(notif['message']).lower(),
                str(notif['telephone']).lower(),
                str(notif['adresse']).lower(),
                str(notif['priorite']).lower()
            ]
            
            # Vérifier si le terme de recherche est dans l'un des champs
              if any(search_term in field for field in searchable_fields):
                   self.notif_tree.insert('', 'end', values=(
                    notif['colis_id'],
                    notif['destinataire'],
                    notif['type'],
                    notif['etat'],
                    notif['date_creation'],
                    notif['priorite']
                ))
                   found = True

        # Si aucun résultat
           if not found:
               messagebox.showinfo("Recherche", "Aucune notification trouvée")
               self.actualiser_notifications()
    

        # Bouton rechercher
        search_button = ctk.CTkButton(
            search_frame,
            text="🔍 Rechercher",
            command=rechercher_notifications,
            fg_color=self.blue_ciel,
            hover_color="#1976D2"
        )
        search_button.pack(side="left", padx=5)

        # Configuration de la recherche en temps réel
        def on_search_change(*args):
            self.rechercher_notifications()
        
        self.search_var.trace_add('write', on_search_change)


        # Filtre
        filter_label = ctk.CTkLabel(search_frame, text="Filtrer par:", text_color="black")
        filter_label.pack(side="left", padx=(20, 5))

        self.filter_var = tk.StringVar(value="Toutes")
        filter_combo = ttk.Combobox(
            search_frame,
            textvariable=self.filter_var,
            values=["Toutes", "Non lues", "Prioritaires", "En attente"],
            width=15,
            state="readonly"
        )
        filter_combo.pack(side="left", padx=5)
        filter_combo.bind('<<ComboboxSelected>>', lambda e: self.actualiser_notifications())

        # Frame pour la table
        table_frame = ctk.CTkFrame(main_container, fg_color="white")
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)

        # En-tête de la table
        columns = ("ID Colis", "Destinataire", "Type", "État", "Date", "Priorité")
        self.notif_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        # Configuration des colonnes
        for col in columns:
            self.notif_tree.heading(col, text=col)
            width = 150 if col in ["Destinataire"] else 100
            self.notif_tree.column(col, width=width)

        # Style pour la table
        style = ttk.Style()
        style.configure("Treeview", rowheight=30, font=('Helvetica', 10))
        style.configure("Treeview.Heading", font=('Helvetica', 10, 'bold'))

        self.notif_tree.pack(fill="both", expand=True)
        self.notif_tree.bind('<Double-Button-1>', self.afficher_details_notification)

        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.notif_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.notif_tree.configure(yscrollcommand=scrollbar.set)

        # Frame pour les boutons d'action
        button_frame = ctk.CTkFrame(main_container, fg_color="white")
        button_frame.pack(fill="x", padx=20, pady=10)

        # Boutons
        ctk.CTkButton(
            button_frame,
            text="✉️ Nouvelle notification",
            command=self.nouvelle_notification,
            fg_color=self.green,
            width=200
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            button_frame,
            text="🗑️ Supprimer la sélection",
            command=self.supprimer_notification,
            fg_color=self.red,
            width=200
        ).pack(side="left", padx=5)

        # Initialisation des données si nécessaire
        if not hasattr(self, 'mock_notifications'):
            self.mock_notifications = []

        # Charger les notifications
        self.actualiser_notifications()

    def afficher_details_notification(self, event):
        """Afficher les détails d'une notification dans une nouvelle fenêtre"""
        selection = self.notif_tree.selection()
        if not selection:
            return

        # Trouver la notification sélectionnée
        item = self.notif_tree.item(selection[0])
        colis_id = item['values'][0]
        notif = next((n for n in self.mock_notifications if n['colis_id'] == colis_id), None)

        if not notif:
            return

        # Créer une nouvelle fenêtre
        details_window = ctk.CTkToplevel(self)
        details_window.title(f"Détails de la notification - {colis_id}")
        details_window.geometry("500x600")

        # Frame principal
        main_frame = ctk.CTkFrame(details_window, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Titre
        title_label = ctk.CTkLabel(
            main_frame,
            text="📋 Détails de la notification",
            font=("Helvetica", 20, "bold"),
            text_color=self.blue_ciel
        )
        title_label.pack(pady=10)

        # Frame pour les détails
        details_frame = ctk.CTkFrame(main_frame, fg_color="white")
        details_frame.pack(fill="both", expand=True, pady=10)

        # Afficher les détails
        details = [
            ("ID Colis", notif['colis_id']),
            ("Destinataire", notif['destinataire']),
            ("Adresse", notif['adresse']),
            ("Téléphone", notif['telephone']),
            ("Montant", f"{notif['montant']} DH"),
            ("État", notif['etat']),
            ("Type", notif['type']),
            ("Message", notif['message']),
            ("Date création", notif['date_creation']),
            ("Livraison prévue", f"{notif['date_livraison_debut']} - {notif['date_livraison_fin']}"),
            ("Priorité", notif['priorite'])
        ]

        for label, value in details:
            detail_frame = ctk.CTkFrame(details_frame, fg_color="transparent")
            detail_frame.pack(fill="x", pady=5)
            
            ctk.CTkLabel(
                detail_frame,
                text=f"{label}:",
                font=("Helvetica", 12, "bold"),
                width=150,
                anchor="e"
            ).pack(side="left", padx=5)
            
            ctk.CTkLabel(
                detail_frame,
                text=str(value),
                font=("Helvetica", 12)
            ).pack(side="left", padx=5)

        # Bouton fermer
        ctk.CTkButton(
            main_frame,
            text="Fermer",
            command=details_window.destroy,
            fg_color=self.red
        ).pack(pady=20)

    def actualiser_notifications(self):
        """Actualiser l'affichage des notifications"""
        # Effacer le tableau
        for item in self.notif_tree.get_children():
            self.notif_tree.delete(item)

        # Réinsérer les données filtrées
        filter_value = self.filter_var.get()
        
        for notif in self.mock_notifications:
            # Appliquer les filtres
            if filter_value == "Non lues" and notif['lu']:
                continue
            if filter_value == "Prioritaires" and notif['priorite'] != "Haute":
                continue
            if filter_value == "En attente" and notif['etat'] != "En attente":
                continue

            # Insérer dans le tableau
            self.notif_tree.insert('', 'end', values=(
                notif['colis_id'],
                notif['destinataire'],
                notif['type'],
                notif['etat'],
                notif['date_creation'],
                notif['priorite']
            ))

        # Si aucune donnée
        if not self.notif_tree.get_children():
            self.notif_tree.insert('', 'end', values=('Aucune notification', '', '', '', '', ''))



        # Ajoutez également un événement de recherche en temps réel (optionnel)
    def setup_search_binding(self):
        def on_search_change(*args):
            self.rechercher_notifications()
        
        self.search_var.trace_add('write', on_search_change)

    def supprimer_notification(self):
        """Supprimer une notification"""
        selection = self.notif_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner une notification")
            return

        if messagebox.askyesno("Confirmation", "Voulez-vous vraiment supprimer cette notification ?"):
            item = self.notif_tree.item(selection[0])
            colis_id = item['values'][0]
            
            # Supprimer des données d'exemple
            self.mock_notifications = [n for n in self.mock_notifications if n['colis_id'] != colis_id]
            
            self.actualiser_notifications()
            messagebox.showinfo("Succès", "Notification supprimée")
    def nouvelle_notification(self):
        """Créer une nouvelle notification"""
        try:
            # Création de la fenêtre modale
            dialog = ctk.CTkToplevel(self)
            dialog.title("Nouvelle Notification")
            dialog.geometry("600x700")
            dialog.transient(self)  # Rend la fenêtre dépendante de la fenêtre principale
            dialog.grab_set()  # Rend la fenêtre modale

            # Configuration de la fenêtre
            dialog.configure(fg_color="white")
            
            # Frame principal
            main_frame = ctk.CTkFrame(dialog, fg_color="white")
            main_frame.pack(fill="both", expand=True, padx=20, pady=20)

            # Titre
            title_label = ctk.CTkLabel(
                main_frame,
                text="✉️ Nouvelle Notification",
                font=("Helvetica", 20, "bold"),
                text_color=self.blue_ciel
            )
            title_label.pack(pady=10)

            # Frame pour le formulaire (utiliser CTkFrame au lieu de CTkScrollableFrame)
            form_frame = ctk.CTkFrame(main_frame, fg_color="white")
            form_frame.pack(fill="both", expand=True, pady=10)

            # Variables de contrôle
            vars_dict = {}
            for field in ["ID Colis", "Destinataire", "Adresse", "Téléphone", "Montant", "Type", "État", "Priorité"]:
                vars_dict[field] = tk.StringVar()

            # Valeurs par défaut
            vars_dict["Type"].set("Livraison")
            vars_dict["État"].set("Nouvelle")
            vars_dict["Priorité"].set("Normale")

            # Création des champs
            for label in ["ID Colis", "Destinataire", "Adresse", "Téléphone", "Montant"]:
                field_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
                field_frame.pack(fill="x", pady=5)
                
                ctk.CTkLabel(
                    field_frame,
                    text=f"{label} :",
                    font=("Helvetica", 12, "bold"),
                    width=120,
                    anchor="e"
                ).pack(side="left", padx=5)
                
                ctk.CTkEntry(
                    field_frame,
                    textvariable=vars_dict[label],
                    width=300
                ).pack(side="left", padx=5, fill="x", expand=True)

            # Comboboxes
            combos = {
                "Type": ["Livraison", "Retard", "Problème", "Information"],
                "État": ["Nouvelle", "En cours", "En attente"],
                "Priorité": ["Normale", "Haute", "Urgente"]
            }

            for label, values in combos.items():
                field_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
                field_frame.pack(fill="x", pady=5)
                
                ctk.CTkLabel(
                    field_frame,
                    text=f"{label} :",
                    font=("Helvetica", 12, "bold"),
                    width=120,
                    anchor="e"
                ).pack(side="left", padx=5)
                
                combo = ttk.Combobox(
                    field_frame,
                    textvariable=vars_dict[label],
                    values=values,
                    state="readonly",
                    width=30
                )
                combo.pack(side="left", padx=5, fill="x", expand=True)
                combo.set(values[0])

            # Message
            message_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
            message_frame.pack(fill="x", pady=5)
            
            ctk.CTkLabel(
                message_frame,
                text="Message :",
                font=("Helvetica", 12, "bold"),
                width=120,
                anchor="e"
            ).pack(side="left", padx=5)
            
            message_text = ctk.CTkTextbox(message_frame, height=100)
            message_text.pack(side="left", padx=5, fill="x", expand=True)

            # Dates de livraison
            dates_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
            dates_frame.pack(fill="x", pady=10)

            ctk.CTkLabel(
                dates_frame,
                text="Livraison :",
                font=("Helvetica", 12, "bold"),
                width=120,
                anchor="e"
            ).pack(side="left", padx=5)

            date_debut = DateEntry(dates_frame, width=12, background=self.blue_ciel, foreground='white')
            date_debut.pack(side="left", padx=5)

            ctk.CTkLabel(dates_frame, text="à").pack(side="left", padx=5)

            date_fin = DateEntry(dates_frame, width=12, background=self.blue_ciel, foreground='white')
            date_fin.pack(side="left", padx=5)

            def sauvegarder():
                try:
                    # Validation du montant
                    try:
                        montant = float(vars_dict["Montant"].get().replace(',', '.')) if vars_dict["Montant"].get().strip() else 0
                        if montant < 0:
                            messagebox.showwarning("Attention", "Le montant ne peut pas être négatif")
                            return
                    except ValueError:
                        messagebox.showwarning("Attention", "Le montant doit être un nombre valide")
                        return

                    # Création de la notification
                    nouvelle_notif = {
                        'id': f'NOT{len(self.mock_notifications) + 1:03d}',
                        'colis_id': vars_dict["ID Colis"].get(),
                        'type': vars_dict["Type"].get(),
                        'message': message_text.get("1.0", "end-1c"),
                        'destinataire': vars_dict["Destinataire"].get(),
                        'adresse': vars_dict["Adresse"].get(),
                        'telephone': vars_dict["Téléphone"].get(),
                        'montant': montant,
                        'etat': vars_dict["État"].get(),
                        'date_creation': datetime.now().strftime('%Y-%m-%d %H:%M'),
                        'date_livraison_debut': date_debut.get_date().strftime('%Y-%m-%d %H:%M'),
                        'date_livraison_fin': date_fin.get_date().strftime('%Y-%m-%d %H:%M'),
                        'priorite': vars_dict["Priorité"].get(),
                        'lu': False
                    }

                    self.mock_notifications.append(nouvelle_notif)
                    self.actualiser_notifications()
                    messagebox.showinfo("Succès", "Notification créée avec succès!")
                    dialog.destroy()

                except Exception as e:
                    messagebox.showerror("Erreur", f"Une erreur est survenue: {str(e)}")

            # Boutons
            button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            button_frame.pack(fill="x", pady=20)

            ctk.CTkButton(
                button_frame,
                text="Annuler",
                command=dialog.destroy,
                fg_color=self.red,
                width=100
            ).pack(side="right", padx=5)

            ctk.CTkButton(
                button_frame,
                text="Sauvegarder",
                command=sauvegarder,
                fg_color=self.green,
                width=100
            ).pack(side="right", padx=5)

            # Focus sur la fenêtre
            dialog.focus_force()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la création de la fenêtre: {str(e)}")
# Exécution de l'application
if __name__ == "__main__":
    root = ctk.CTk()
    root.geometry("1200x800")
    app = AdminPage(root)
    app.pack(fill="both", expand=True)
    root.mainloop()









